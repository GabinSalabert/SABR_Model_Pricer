from SABR_function import *
from input import*
from openpyxl import workbook
from openpyxl.cell import get_column_letter

time, tenor = maturite(dfvols)
strike = swaprates
MKT = vol(dfvols)
spot = swaprates

alpha, beta, rho, nu = calibration(starting_guess, f, strike, time, MKT)
matrix_vol = SABR_vol_matrix(alpha, beta, rho, nu, f, strike, time)
prices_call = bachelier_matrix(matrix_vol, spot, strike, f, time, "call")
prices_put = bachelier_matrix(matrix_vol, spot, strike, f, time, "put")

wb = openpyxl.load_workbook('Option_3_DataSet.xls')
sheet = wb.get_sheet_by_name('Option_3_DataSet')
for data in enumerate(prices_call):
    worksheet.write_column(2, 9, data)
for data in enumerate(prices_put):
    worksheet.write_column(2, 10, data)
for data in enumerate(time):
    worksheet.write_column(2, 11, data)
for data in enumerate(tenor):
    worksheet.write_column(2, 12, data)
for data in enumerate(strike):
    worksheet.write_column(2, 13, data)
for data in enumerate(MKT):
    worksheet.write_column(2, 14, data)
for data in enumerate(spot):
    worksheet.write_column(2, 15, data)
    
sheet.cell(row=(2),column=5).value = alpha
sheet.cell(row=(2),column=6).value = beta
sheet.cell(row=(2),column=7).value = rho
sheet.cell(row=(2),column=8).value = nu

wb.save('Option_3_DataSet.xls')
